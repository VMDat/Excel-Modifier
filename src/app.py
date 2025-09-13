import pandas as pd
import streamlit as st
from openpyxl import Workbook, load_workbook
import os
import io

def process_files(uploaded_files, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Tổng hợp"

    # Xử lý từ từng file đã upload
    for uploaded_file in uploaded_files:
        try:
            filename = uploaded_file.name
            base_name = os.path.splitext(filename)[0]
            
            if len(base_name) != 8:
                st.warning(f"Skipped {filename}: Invalid name format (should be 8 characters)")
                continue
            
            xx_str = base_name[0:2]
            yy_str = base_name[4:6]
            zz_str = base_name[6:8]
                
            try:
                xx = int(xx_str)
            except ValueError:
                st.warning(f"Skipped {filename}: XX must be numeric")
                continue
            
            row_in_excel = xx + 6
            
            # Xác định cột ghi kết quả
            if yy_str == '51':
                if zz_str == '01': col_offset = 4
                elif zz_str == '02': continue
                elif zz_str == '03': col_offset = 6
                else: continue
            elif yy_str == '53':
                if zz_str == '01': col_offset = 8
                elif zz_str == '02': col_offset = 10
                elif zz_str == '03': col_offset = 12
                elif zz_str == '04': col_offset = 14
                elif zz_str == '05': col_offset = 16
                elif zz_str == '06': col_offset = 18
                else: continue
            else:
                continue
            
            # Đọc và xử lý CSV
            df = pd.read_csv(uploaded_file, header=None, nrows=2)
            sum_row0 = df.iloc[0, 2:50].sum()
            sum_row1 = df.iloc[1, 2:50].sum()
            
            # Ghi kết quả vào Excel
            ws.cell(row=row_in_excel, column=col_offset, value=sum_row0)
            ws.cell(row=row_in_excel, column=col_offset+1, value=sum_row1)
            
            st.success(f"Processed {filename} → Row {row_in_excel}, Columns {col_offset}-{col_offset+1}")
            
        except Exception as e:
            st.error(f"Error processing {filename}: {str(e)}")
    
    # Lưu file Excel
    wb.save(output_file)
    return output_file

def main():
    st.title("CSV Processing Tool")
    st.subheader("Process CSV files and export to Excel")
    
    # Upload file
    uploaded_files = st.file_uploader(
        "Select CSV files", 
        type=['csv', 'CSV'], 
        accept_multiple_files=True
    )
    
    # Chọn file output
    output_file = st.text_input("Output Excel file name", "output.xlsx")
    
    if st.button("Process Files"):
        if not uploaded_files:
            st.warning("Please upload at least one CSV file")
            return
            
        with st.spinner("Processing files..."):
            try:
                # Xử lý file
                output_path = process_files(uploaded_files, output_file)
                
                # Hiển thị link download
                with open(output_path, "rb") as f:
                    st.download_button(
                        label="Download Result",
                        data=f,
                        file_name=output_file,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                st.success("Processing completed!")
                
            except Exception as e:
                st.error(f"Error: {str(e)}")

if __name__ == "__main__":
    main()